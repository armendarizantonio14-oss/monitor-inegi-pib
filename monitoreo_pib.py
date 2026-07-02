import requests
import json
import os
import smtplib

from email.mime.text import MIMEText
from datetime import datetime
from openpyxl import Workbook, load_workbook

# =====================================================
# CONFIGURACIÓN
# =====================================================

URL = "https://www.inegi.org.mx/app/tabulados/serviciocuadros/wsDataService.svc/obtienedatotabturistatp2/6207137914,6207137915,6207137859/si/2/pv3/1"

ARCHIVO_CONTROL = "historial_pib.json"
ARCHIVO_EXCEL = "monitoreo_pib.xlsx"

# =====================================================
# CONVERTIR PERIODO
# =====================================================

def convertir_periodo(periodo):

    equivalencias = {
        "I": "1T",
        "II": "2T",
        "III": "3T",
        "IV": "4T",
        "VI": "6 meses",
        "IX": "9 meses",
        "XII": "Anual"
    }

    return equivalencias.get(periodo, periodo)

# =====================================================
# CORREO
# =====================================================
EMAIL_ORIGEN = os.getenv("EMAIL_REMITENTE")
PASSWORD = os.getenv("EMAIL_PASSWORD")

EMAIL_DESTINOS = os.getenv(
    "EMAIL_DESTINATARIOS",
    ""
).split(",")
# =====================================================
# EXCEL
# =====================================================

def inicializar_excel():

    if os.path.exists(ARCHIVO_EXCEL):
        return

    wb = Workbook()

    # -----------------------------------------
    # Hoja cambios
    # -----------------------------------------

    ws1 = wb.active
    ws1.title = "Cambios_Historicos"

    ws1.append([
        "Fecha",
        "Año",
        "Indicador",
        "Tipo Cambio",
        "Valor Anterior",
        "Valor Actual",
        "Diferencia",
        "Variacion (%)"
    ])

    # -----------------------------------------
    # Hoja ejecuciones
    # -----------------------------------------

    ws2 = wb.create_sheet("Control_Ejecuciones")

    ws2.append([
        "Fecha",
        "Resultado"
    ])

    # -----------------------------------------
    # Hoja estado actual
    # -----------------------------------------

    ws3 = wb.create_sheet("Estado_Actual")

    ws3.append([
        "Año",
        "I",
        "II",
        "III",
        "IV",
        "VI",
        "IX",
        "XII"
    ])

    wb.save(ARCHIVO_EXCEL)

def actualizar_estado_actual(snapshot):

    wb = load_workbook(ARCHIVO_EXCEL)

    ws = wb["Estado_Actual"]

    ws.delete_rows(1, ws.max_row)

    ws.append([
        "Año",
        "I",
        "II",
        "III",
        "IV",
        "VI",
        "IX",
        "XII"
    ])

    for anio in sorted(snapshot.keys()):

        fila = snapshot[anio]

        ws.append([
            anio,
            fila.get("I", ""),
            fila.get("II", ""),
            fila.get("III", ""),
            fila.get("IV", ""),
            fila.get("VI", ""),
            fila.get("IX", ""),
            fila.get("XII", "")
        ])

    wb.save(ARCHIVO_EXCEL)

def registrar_cambios_excel(cambios):

    wb = load_workbook(ARCHIVO_EXCEL)

    ws = wb["Cambios_Historicos"]

    fecha = datetime.now().strftime(
        "%Y-%m-%d %H:%M:%S"
    )

    for cambio in cambios:

        diferencia = None
        variacion = None

        try:

            if (
                cambio["anterior"] not in ["", None]
                and cambio["nuevo"] not in ["", None]
            ):

                anterior = float(
                    cambio["anterior"]
                )

                nuevo = float(
                    cambio["nuevo"]
                )

                diferencia = (
                    nuevo - anterior
                )

                if anterior != 0:

                    variacion = (
                        diferencia / anterior
                    ) * 100

        except:
            pass

        ws.append([
            fecha,
            cambio["anio"],
            cambio["campo"],
            cambio["tipo"],
            cambio["anterior"],
            cambio["nuevo"],
            diferencia,
            variacion
        ])

    wb.save(ARCHIVO_EXCEL)

def registrar_ejecucion(resultado):

    wb = load_workbook(ARCHIVO_EXCEL)

    ws = wb["Control_Ejecuciones"]

    fecha = datetime.now().strftime(
        "%Y-%m-%d %H:%M:%S"
    )

    ws.append([
        fecha,
        resultado
    ])

    wb.save(ARCHIVO_EXCEL)

# =====================================================
# OBTENER DATOS
# =====================================================

def obtener_datos():

    response = requests.get(
        URL,
        timeout=30
    )

    response.raise_for_status()

    return response.json()["Data"]

# =====================================================
# SNAPSHOT
# =====================================================

def construir_snapshot(data):

    ultimo_anio = data[-1]["C0"]

    snapshot = {}

    for fila in data:

        anio = fila["C0"]

        # -----------------------------
        # Año actual
        # -----------------------------

        if anio == ultimo_anio:

            snapshot[anio] = {

                "I": fila.get("I", ""),

                "II": fila.get("II", ""),

                "III": fila.get("III", ""),

                "IV": fila.get("IV", ""),

                "VI": fila.get("VI", ""),

                "IX": fila.get("IX", ""),

                "XII": fila.get("XII", "")
            }

        # -----------------------------
        # Históricos
        # -----------------------------

        elif int(anio) >= int(ultimo_anio) - 4:

            snapshot[anio] = {

                "XII": fila.get(
                    "XII",
                    ""
                )

            }

    return snapshot

# =====================================================
# JSON
# =====================================================

def guardar_snapshot(snapshot):

    with open(
        ARCHIVO_CONTROL,
        "w",
        encoding="utf-8"
    ) as f:

        json.dump(
            snapshot,
            f,
            indent=4,
            ensure_ascii=False
        )

def cargar_snapshot():

    if not os.path.exists(
        ARCHIVO_CONTROL
    ):
        return None

    with open(
        ARCHIVO_CONTROL,
        "r",
        encoding="utf-8"
    ) as f:

        return json.load(f)

# =====================================================
# COMPARAR
# =====================================================

def comparar(actual, anterior):

    cambios = []

    for anio in actual:

        if anio not in anterior:
            continue

        for campo in actual[anio]:

            valor_actual = actual[anio][campo]

            valor_anterior = anterior[anio].get(
                campo,
                ""
            )

            if valor_actual != valor_anterior:

                tipo = "Corrección"

                if (
                    valor_anterior in [
                        "",
                        None
                    ]
                    and
                    valor_actual not in [
                        "",
                        None
                    ]
                ):

                    tipo = (
                        "Nuevo dato publicado"
                    )

                cambios.append({

                    "anio": anio,

                    "campo": convertir_periodo(campo),

                    "anterior":
                        valor_anterior,

                    "nuevo":
                        valor_actual,

                    "tipo":
                        tipo
                })

    return cambios

# =====================================================
# REPORTE
# =====================================================

def mostrar_cambios(cambios):

    if not cambios:

        print(
            "\n✅ Sin cambios detectados"
        )

        return

    print("\n")
    print("=" * 60)
    print("CAMBIOS DETECTADOS")
    print("=" * 60)

    for c in cambios:

        print()

        print(
            f"Año: {c['anio']}"
        )

        print(
            f"Indicador: {c['campo']}"
        )

        print(
            f"Tipo: {c['tipo']}"
        )

        print(
            f"Anterior: {c['anterior']}"
        )

        print(
            f"Nuevo: {c['nuevo']}"
        )

# =====================================================
# CORREO
# =====================================================

def enviar_correo(cambios):

    if not cambios:
        return

    asunto = (
        f"ALERTA PIB INEGI - "
        f"{len(cambios)} cambio(s)"
    )

    cuerpo = []

    cuerpo.append(
        "Se detectaron cambios en la información del PIB.\n"
    )

    cuerpo.append(
        f"Fecha: {datetime.now()}\n"
    )

    cuerpo.append(
        "-" * 50 + "\n"
    )

    for c in cambios:

        cuerpo.append(
            f"Año: {c['anio']}"
        )

        cuerpo.append(
            f" | Indicador: {c['campo']}"
        )

        cuerpo.append(
            f" | Tipo: {c['tipo']}\n"
        )

        cuerpo.append(
            f"Anterior: {c['anterior']}\n"
        )

        cuerpo.append(
            f"Nuevo: {c['nuevo']}\n"
        )

        cuerpo.append(
            "-" * 50 + "\n"
        )

    mensaje = MIMEText(
        "".join(cuerpo),
        "plain",
        "utf-8"
    )

    mensaje["Subject"] = asunto
    mensaje["From"] = EMAIL_ORIGEN
    mensaje["To"] = ", ".join(EMAIL_DESTINOS)

    try:
        servidor = smtplib.SMTP(
            "smtp.gmail.com",
            587
        )

        servidor.starttls()
        servidor.login(
            EMAIL_ORIGEN,
            PASSWORD
        )

        servidor.send_message(
            mensaje,
            from_addr=EMAIL_ORIGEN,
            to_addrs=EMAIL_DESTINOS
        )

        servidor.quit()

        print(
            "\n📧 Correo enviado correctamente"
        )

    except Exception as e:

        print(
            f"\n❌ Error enviando correo: {e}"
        )

# =====================================================
# MAIN
# =====================================================

def main():

    print(
        "\nConsultando INEGI..."
    )

    inicializar_excel()

    data = obtener_datos()

    snapshot_actual = (
        construir_snapshot(data)
    )

    snapshot_anterior = (
        cargar_snapshot()
    )

    # ---------------------------------
    # Primera ejecución
    # ---------------------------------

    if snapshot_anterior is None:

        guardar_snapshot(
            snapshot_actual
        )

        actualizar_estado_actual(
            snapshot_actual
        )

        registrar_ejecucion(
            "Primera ejecución"
        )

        print(
            "\nPrimera ejecución."
        )

        print(
            "Estado inicial guardado."
        )

        return

    # ---------------------------------
    # Comparar
    # ---------------------------------

    cambios = comparar(
        snapshot_actual,
        snapshot_anterior
    )

    mostrar_cambios(
        cambios
    )

    if cambios:
        registrar_cambios_excel(
            cambios
        )

        enviar_correo(
            cambios
        )

        registrar_ejecucion(
            f"Cambios detectados ({len(cambios)})"
        )

    else:

        registrar_ejecucion(
            "Sin cambios"
        )

    actualizar_estado_actual(
        snapshot_actual
    )

    guardar_snapshot(
        snapshot_actual
    )

    print(
        "\nProceso finalizado."
    )

# =====================================================
# EJECUTAR
# =====================================================

if __name__ == "__main__":

    main()
